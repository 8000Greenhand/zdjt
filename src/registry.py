from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


HEADERS = [
    "时间",
    "商品标题",
    "截图路径",
    "商品URL",
    "页面标题",
    "备注",
]


def append_registry(registry_path: str | Path, row: dict[str, Any]) -> None:
    path = Path(registry_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "截图进度"
        ws.append(HEADERS)
        ws.freeze_panes = "A2"

    ws.append([
        row.get("时间") or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        row.get("商品标题", ""),
        row.get("截图路径", ""),
        row.get("商品URL", ""),
        row.get("页面标题", ""),
        row.get("备注", ""),
    ])

    wb.save(path)
